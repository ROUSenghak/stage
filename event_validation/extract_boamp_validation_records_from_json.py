"""Extract BOAMP raw records needed for the 150-row manual validation sample.

The raw BOAMP export is 465 JSON files (~513 MB) across two year-partitioned
directories. This script extracts only the records whose idweb appears in the
validation sample, using year-prefix filtering to minimise I/O.

Usage (from project root):
  python event_validation/extract_boamp_validation_records_from_json.py
  python event_validation/extract_boamp_validation_records_from_json.py --raw-json-dir data/raw/boamp_sample
  python event_validation/extract_boamp_validation_records_from_json.py --raw-json path/to/export.json

Outputs (event_validation/outputs/):
  boamp_150_sample_records_from_raw_json.csv
  boamp_150_sample_missing_ids_from_raw_json.csv
  boamp_event_validation_audit.xlsx  (BOAMP_Raw_Records sheet appended)
  ../boamp_raw_record_extraction_note.md
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
SAMPLE_CSV = ROOT / "event_validation" / "outputs" / "manual_validation_sample.csv"
OUT_DIR = ROOT / "event_validation" / "outputs"
XLSX_PATH = OUT_DIR / "boamp_event_validation_audit.xlsx"
NOTE_PATH = ROOT / "event_validation" / "boamp_raw_record_extraction_note.md"

BOAMP_BASE_URL = "https://www.boamp.fr/avis/detail/"

# Candidate raw-JSON directories in preference order
RAW_JSON_CANDIDATES = [
    ROOT / "data" / "raw" / "boamp_full",
    ROOT / "data" / "raw" / "boamp_sample",
]

# Import normalisation helpers from scripts/utils.py
sys.path.insert(0, str(ROOT / "scripts"))
try:
    from utils import (
        parse_donnees,
        extract_cpv_codes,
        extract_duration_months,
        extract_buyer_siret,
        extract_amount_eur,
        is_eforms,
        _scalar,
        deep_get,
    )
    _HAS_UTILS = True
except ImportError:
    _HAS_UTILS = False
    print("WARNING: scripts/utils.py not importable — CPV/duration/amount extraction disabled.")


# ---------------------------------------------------------------------------
# 1. CLI
# ---------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--raw-json", metavar="FILE", help="Single raw BOAMP JSON file.")
    p.add_argument("--raw-json-dir", metavar="DIR", help="Directory of JSON files.")
    return p.parse_args()


# ---------------------------------------------------------------------------
# 2. Collect target idwebs from validation sample
# ---------------------------------------------------------------------------
def collect_target_ids(sample: pd.DataFrame) -> dict[str, set[str]]:
    """Return {idweb: set_of_roles} for all ids referenced in the sample."""
    id_roles: dict[str, set[str]] = {}

    def _add(id_str: str, role: str) -> None:
        iw = str(id_str).replace("BOAMP:", "").strip()
        if iw and iw != "nan":
            id_roles.setdefault(iw, set()).add(role)

    for val in sample["contract_id"].dropna():
        _add(val, "source")

    for val in sample["renewal_contract_id"].dropna():
        _add(val, "candidate")

    # nearest_later_notice_id — already stripped (no BOAMP: prefix)
    nearest_col = "nearest_later_notice_id"
    if nearest_col in sample.columns:
        for val in sample[nearest_col].dropna():
            _add(val, "nearest_later")

    return id_roles


# ---------------------------------------------------------------------------
# 3. Year-based file discovery
# ---------------------------------------------------------------------------
def _year_from_idweb(idweb: str) -> int | None:
    """'15-30334' → 2015, '24-1234' → 2024."""
    m = re.match(r"^(\d{2})-", idweb)
    if not m:
        return None
    yy = int(m.group(1))
    return 2000 + yy


def discover_json_files(raw_dir: Path, needed_years: set[int]) -> list[Path]:
    """Return JSON files in `raw_dir` whose filename year matches needed_years."""
    files = []
    for f in sorted(raw_dir.glob("*.json")):
        # Match patterns: "2015_p1.json" or "boamp_2015_p1.json"
        m = re.search(r"(\d{4})_p\d+\.json$", f.name)
        if m:
            year = int(m.group(1))
            if year in needed_years:
                files.append(f)
    return files


# ---------------------------------------------------------------------------
# 4. Parse a single JSON file (one BOAMP page export)
# ---------------------------------------------------------------------------
def parse_json_file(path: Path) -> list[dict]:
    """Return list of record dicts from one JSON export file."""
    try:
        with open(path, encoding="utf-8") as fh:
            content = fh.read().strip()
    except OSError as e:
        print(f"  WARN: cannot read {path}: {e}")
        return []

    if not content:
        return []

    # Try standard JSON first
    try:
        data = json.loads(content)
        if isinstance(data, list):
            return data  # bare JSON array
        if isinstance(data, dict):
            # Opendatasoft export: {"total_count": N, "results": [...]}
            records = data.get("results") or data.get("records") or []
            # Handle Opendatasoft v1 format: {"records": [{"fields": {...}}]}
            out = []
            for r in records:
                if isinstance(r, dict) and "fields" in r:
                    out.append(r["fields"])
                else:
                    out.append(r)
            return out
    except json.JSONDecodeError:
        pass

    # Fallback: newline-delimited JSON (NDJSON)
    records = []
    for line in content.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError:
            pass
    return records


# ---------------------------------------------------------------------------
# 5. Normalise one extracted record
# ---------------------------------------------------------------------------
def _joinlist(value) -> str:
    if isinstance(value, list):
        return "|".join(str(v) for v in value)
    return str(value) if value is not None else ""


def normalise_record(record: dict, roles: set[str]) -> dict:
    """Flatten a raw BOAMP record into normalised output columns."""
    idweb = str(record.get("idweb", "")).strip()

    # Parse donnees once
    if _HAS_UTILS:
        donnees = parse_donnees(record)
        cpv_codes = extract_cpv_codes(donnees)
        duration, _ = extract_duration_months(donnees)
        buyer_siret = extract_buyer_siret(donnees)
        amount = extract_amount_eur(donnees)
        donnees_format = "eforms" if is_eforms(donnees) else "legacy"
        cpv_principal = (
            _scalar(deep_get(donnees, "OBJET.CPV.PRINCIPAL"))
            or (cpv_codes[0] if cpv_codes else None)
        )
    else:
        cpv_codes, duration, buyer_siret, amount = [], None, None, None
        donnees_format = "unknown"
        cpv_principal = None

    raw_url = record.get("url_avis") or ""
    if not raw_url:
        raw_url = BOAMP_BASE_URL + idweb

    return {
        "idweb": idweb,
        "role": "|".join(sorted(roles)),
        "dateparution": record.get("dateparution", ""),
        "nomacheteur": record.get("nomacheteur", ""),
        "buyer_siret": buyer_siret or "",
        "objet": record.get("objet", ""),
        "cpv_principal": cpv_principal or "",
        "cpv_all": "|".join(cpv_codes),
        "type_avis": _joinlist(record.get("type_avis")),
        "type_procedure": record.get("type_procedure", ""),
        "type_marche": _joinlist(record.get("type_marche")),
        "amount_eur": amount,
        "duration_months": duration,
        "url_avis": raw_url,
        "donnees_format": donnees_format,
        "raw_record_json": json.dumps(record, ensure_ascii=False),
    }


# ---------------------------------------------------------------------------
# 6. Main extraction logic
# ---------------------------------------------------------------------------
def extract_records(
    source: Path | None,
    source_dir: Path | None,
    target_id_roles: dict[str, set[str]],
) -> tuple[list[dict], set[str]]:
    """Scan JSON source(s) and return (matched_normalised_records, found_ids)."""
    target_ids = set(target_id_roles.keys())
    found: dict[str, dict] = {}  # idweb → normalised record

    def _scan_records(records: list[dict]) -> None:
        for rec in records:
            iw = str(rec.get("idweb", "")).strip()
            if iw and iw in target_ids and iw not in found:
                roles = target_id_roles[iw]
                found[iw] = normalise_record(rec, roles)

    if source:
        print(f"  Scanning single file: {source}")
        _scan_records(parse_json_file(source))
    elif source_dir:
        needed_years = set()
        for iw in target_ids:
            y = _year_from_idweb(iw)
            if y:
                needed_years.add(y)
        files = discover_json_files(source_dir, needed_years)
        print(f"  Needed years: {sorted(needed_years)}")
        print(f"  Files to scan: {len(files)} (out of {len(list(source_dir.glob('*.json')))} total in dir)")
        for f in files:
            recs = parse_json_file(f)
            _scan_records(recs)
            # Early exit if all found
            if len(found) == len(target_ids):
                break

    return list(found.values()), set(found.keys())


# ---------------------------------------------------------------------------
# 7. Update Excel workbook — append sheet without overwriting
# ---------------------------------------------------------------------------
def append_excel_sheet(xlsx_path: Path, df_excel: pd.DataFrame) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    sheet_name = "BOAMP_Raw_Records"
    wb = load_workbook(xlsx_path)

    # Remove sheet if already present (re-run idempotency)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    rows = list(dataframe_to_rows(df_excel, index=False, header=True))
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    # Auto-size columns (cap at 50)
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(xlsx_path)
    print(f"  Sheet '{sheet_name}' appended to {xlsx_path.name}")


# ---------------------------------------------------------------------------
# 8. Markdown note
# ---------------------------------------------------------------------------
def write_markdown_note(
    raw_source: str,
    n_files_scanned: int,
    n_requested: int,
    n_found: int,
    n_missing: int,
) -> None:
    content = f"""# BOAMP Raw Record Extraction — Note

## Why this file exists

The full BOAMP open-data export for Pays de la Loire (2015–2024, digital CPV divisions) is
stored locally as ~465 JSON files totalling approximately 513 MB across two directories:

- `data/raw/boamp_full/` — 348 files, ~384 MB (authoritative full export)
- `data/raw/boamp_sample/` — 117 files, ~129 MB (subset used for development)

This file is too large to include in a shared repository or to process in full for a
150-row validation exercise. Instead, the script
`event_validation/extract_boamp_validation_records_from_json.py`
extracted only the **{n_found} records** (out of {n_requested} requested) whose
`idweb` appears in the manual validation sample.

## What was extracted

| Item | Count |
|---|---|
| Unique idwebs requested (source + candidate + nearest-later) | {n_requested} |
| Records found in JSON export | {n_found} |
| Missing IDs (not found in local export) | {n_missing} |
| Raw JSON source used | `{raw_source}` |
| Files scanned | {n_files_scanned} |

Missing IDs are documented in:
`event_validation/outputs/boamp_150_sample_missing_ids_from_raw_json.csv`

They may be absent because:
- The record was published outside the fetched year range or CPV pre-filter.
- The notice was removed or superseded on the BOAMP platform.
- The `idweb` was assigned by the matching algorithm from a different data vintage.

## How to re-run

```bash
cd /path/to/stage-1
python event_validation/extract_boamp_validation_records_from_json.py
# Or specify a different directory:
python event_validation/extract_boamp_validation_records_from_json.py --raw-json-dir data/raw/boamp_sample
```

## Output files

- `event_validation/outputs/boamp_150_sample_records_from_raw_json.csv` — one row per extracted record
- `event_validation/outputs/boamp_150_sample_missing_ids_from_raw_json.csv` — unmatched IDs
- `event_validation/outputs/boamp_event_validation_audit.xlsx` — sheet `BOAMP_Raw_Records` appended
"""
    NOTE_PATH.write_text(content, encoding="utf-8")
    print(f"  Markdown note written: {NOTE_PATH.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    args = parse_args()

    # --- Resolve JSON source ---
    single_file: Path | None = None
    json_dir: Path | None = None

    if args.raw_json:
        single_file = Path(args.raw_json).resolve()
        if not single_file.is_file():
            sys.exit(f"ERROR: --raw-json file not found: {single_file}")
        print(f"Mode: single file — {single_file}")
    elif args.raw_json_dir:
        json_dir = Path(args.raw_json_dir).resolve()
        if not json_dir.is_dir():
            sys.exit(f"ERROR: --raw-json-dir not found: {json_dir}")
        print(f"Mode: directory — {json_dir}")
    else:
        # Auto-detect
        print("Auto-detecting raw JSON source …")
        for candidate in RAW_JSON_CANDIDATES:
            print(f"  Checking: {candidate} — ", end="")
            if candidate.is_dir():
                n = len(list(candidate.glob("*.json")))
                print(f"found ({n} JSON files)")
                json_dir = candidate
                break
            else:
                print("not found")
        if json_dir is None:
            sys.exit(
                "ERROR: No raw JSON directory found. Use --raw-json or --raw-json-dir."
            )

    print(f"\nRaw JSON source: {single_file or json_dir}")

    # --- Load validation sample ---
    if not SAMPLE_CSV.exists():
        sys.exit(f"ERROR: validation sample not found: {SAMPLE_CSV}")
    sample = pd.read_csv(SAMPLE_CSV, dtype=str)
    print(f"\nValidation sample: {len(sample)} rows")

    # --- Collect target IDs ---
    id_roles = collect_target_ids(sample)
    print(f"Unique target idwebs: {len(id_roles)}")

    # --- Extract ---
    print("\nScanning JSON files …")
    matched_records, found_ids = extract_records(single_file, json_dir, id_roles)

    # --- Missing IDs ---
    missing = {iw: id_roles[iw] for iw in id_roles if iw not in found_ids}

    print(f"\n--- Extraction summary ---")
    print(f"  Target IDs:        {len(id_roles)}")
    print(f"  Records extracted: {len(matched_records)}")
    print(f"  Missing IDs:       {len(missing)}")

    # --- Save extracted records CSV ---
    if matched_records:
        df_records = pd.DataFrame(matched_records)
        # Sort: sources first, then candidates, then nearest_later
        role_order = {"source": 0, "candidate": 1, "nearest_later": 2}
        df_records["_sort"] = df_records["role"].map(
            lambda r: min(role_order.get(x, 9) for x in r.split("|"))
        )
        df_records = df_records.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)
    else:
        df_records = pd.DataFrame(
            columns=["idweb", "role", "dateparution", "nomacheteur", "buyer_siret",
                     "objet", "cpv_principal", "cpv_all", "type_avis", "type_procedure",
                     "type_marche", "amount_eur", "duration_months", "url_avis",
                     "donnees_format", "raw_record_json"]
        )

    out_records = OUT_DIR / "boamp_150_sample_records_from_raw_json.csv"
    df_records.to_csv(out_records, index=False)
    print(f"\n  CSV written: {out_records.name} ({len(df_records)} rows)")

    # --- Save missing IDs CSV ---
    missing_rows = [
        {"idweb": iw, "role": "|".join(sorted(roles))}
        for iw, roles in sorted(missing.items())
    ]
    df_missing = pd.DataFrame(missing_rows) if missing_rows else pd.DataFrame(
        columns=["idweb", "role"]
    )
    out_missing = OUT_DIR / "boamp_150_sample_missing_ids_from_raw_json.csv"
    df_missing.to_csv(out_missing, index=False)
    print(f"  CSV written: {out_missing.name} ({len(df_missing)} rows)")

    # --- Update Excel (skip raw_record_json — too wide) ---
    if XLSX_PATH.exists():
        excel_cols = [c for c in df_records.columns if c != "raw_record_json"]
        try:
            append_excel_sheet(XLSX_PATH, df_records[excel_cols])
        except Exception as e:
            print(f"  WARN: Excel update failed: {e}")
    else:
        print(f"  WARN: Excel file not found at {XLSX_PATH} — skipping sheet append.")

    # --- Markdown note ---
    n_files_scanned = (
        len(list(json_dir.glob("*.json"))) if json_dir else 1
    )
    write_markdown_note(
        raw_source=str(single_file or json_dir),
        n_files_scanned=n_files_scanned,
        n_requested=len(id_roles),
        n_found=len(matched_records),
        n_missing=len(missing),
    )

    print(f"\nAll outputs written to: {OUT_DIR}")


if __name__ == "__main__":
    main()
