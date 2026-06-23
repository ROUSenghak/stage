"""Build the 150-row manual validation sample for the BOAMP renewal proxy event.

Outputs (all in event_validation/outputs/):
  manual_validation_sample.csv
  validation_summary_metrics.csv
  threshold_sensitivity.csv
  event_bias_summary.csv
  event_rate_by_category.csv
  boamp_event_validation_audit.xlsx

Run from the project root:
  python event_validation/scripts/build_validation_sample.py
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent.parent
SURVIVAL_PATH = ROOT / "data" / "processed" / "boamp_phase2_survival.csv"
CANDIDATES_PATH = (
    ROOT / "boamp_renewal_linking_quality" / "outputs" / "boamp_renewal_candidates.csv"
)
FULL_CLEAN_PATH = ROOT / "data" / "processed" / "boamp_full_clean.csv"
OUT_DIR = ROOT / "event_validation" / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)

RANDOM_SEED = 42
TARGET_N = 150
BOAMP_BASE_URL = "https://www.boamp.fr/avis/detail/"

# Thresholds for sensitivity table
SCORE_THRESHOLDS = [0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80]

# Allowed manual_decision values (kept as comment for auditor reference)
MANUAL_DECISION_VALUES = [
    "credible_renewal",
    "doubtful_but_possible",
    "not_credible_false_positive",
    "plausible_censored",
    "missed_renewal_false_negative",
    "impossible_to_judge",
]
MANUAL_ERROR_TYPE_VALUES = [
    "true_positive",
    "false_positive",
    "true_negative_or_plausible_censored",
    "false_negative",
    "not_applicable",
]


# ---------------------------------------------------------------------------
# 1. Load data
# ---------------------------------------------------------------------------
def load_survival() -> pd.DataFrame:
    df = pd.read_csv(
        SURVIVAL_PATH,
        dtype={"contract_id": str, "renewal_contract_id": str, "buyer_key": str},
        parse_dates=["start_date", "estimated_end_date"],
        low_memory=False,
    )
    df["event"] = pd.to_numeric(df["event"], errors="coerce").fillna(0).astype(int)
    df["composite_score"] = pd.to_numeric(df["composite_score"], errors="coerce")
    df["score_margin"] = pd.to_numeric(df["score_margin"], errors="coerce")
    df["n_candidates_for_source"] = pd.to_numeric(
        df["n_candidates_for_source"], errors="coerce"
    )
    df["single_candidate_match"] = (
        df["single_candidate_match"].astype(str).str.lower().isin(["true", "1"])
    )
    df["high_confidence_strict"] = (
        df["high_confidence_strict"].astype(str).str.lower().isin(["true", "1"])
    )
    df["declared_duration_months"] = pd.to_numeric(
        df["declared_duration_months"], errors="coerce"
    )
    df["observed_duration_months"] = pd.to_numeric(
        df["observed_duration_months"], errors="coerce"
    )
    df["text_similarity"] = pd.to_numeric(df["text_similarity"], errors="coerce")
    df["cpv_match_score"] = pd.to_numeric(df["cpv_match_score"], errors="coerce")
    df["temporal_score"] = pd.to_numeric(df["temporal_score"], errors="coerce")
    return df


def load_candidates() -> pd.DataFrame:
    df = pd.read_csv(
        CANDIDATES_PATH,
        dtype={
            "src_idweb": str,
            "cand_idweb": str,
            "src_buyer_key": str,
            "src_cpv": str,
            "cand_cpv": str,
        },
        low_memory=False,
    )
    df["composite_score"] = pd.to_numeric(df["composite_score"], errors="coerce")
    # Keep only the best-scoring candidate per source contract
    df = df.sort_values("composite_score", ascending=False)
    df = df.drop_duplicates(subset=["src_idweb"], keep="first").copy()
    return df.set_index("src_idweb")


def load_full_clean() -> pd.DataFrame:
    usecols = ["idweb", "nomacheteur", "objet", "url_avis", "buyer_key", "dateparution"]
    df = pd.read_csv(
        FULL_CLEAN_PATH,
        dtype={"idweb": str, "buyer_key": str},
        usecols=usecols,
        low_memory=False,
        parse_dates=["dateparution"],
    )
    return df


# ---------------------------------------------------------------------------
# 2. Feature engineering for stratification
# ---------------------------------------------------------------------------
def add_strata_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["start_year"] = df["start_date"].dt.year.fillna(0).astype(int)

    def margin_bin(row):
        if row["event"] != 1 or pd.isna(row["score_margin"]):
            return None
        if row["score_margin"] < 0.05:
            return "low"
        if row["score_margin"] < 0.15:
            return "medium"
        return "high"

    df["margin_bin"] = df.apply(margin_bin, axis=1)
    return df


# ---------------------------------------------------------------------------
# 3. Stratified 150-row sample
# ---------------------------------------------------------------------------
def stratified_sample(df: pd.DataFrame, rng: np.random.Generator) -> pd.DataFrame:
    sampled_ids: set[str] = set()
    frames: list[pd.DataFrame] = []

    def _pick(pool: pd.DataFrame, n: int, secondary_cols: list[str] | None = None) -> pd.DataFrame:
        pool = pool[~pool["contract_id"].isin(sampled_ids)].copy()
        if pool.empty or n <= 0:
            return pd.DataFrame(columns=pool.columns)
        if secondary_cols and len(pool) > n:
            # Proportional allocation across secondary strata
            pool["_stratum"] = pool[secondary_cols].astype(str).agg("-".join, axis=1)
            counts = pool["_stratum"].value_counts(normalize=True)
            picked = []
            budget = n
            strata_keys = list(counts.index)
            for key in strata_keys:
                quota = max(1, round(counts[key] * n))
                sub = pool[pool["_stratum"] == key]
                take = min(quota, len(sub), budget)
                if take > 0:
                    idx = rng.choice(len(sub), size=take, replace=False)
                    picked.append(sub.iloc[idx])
                    budget -= take
                if budget <= 0:
                    break
            if budget > 0:
                remaining = pool[~pool["contract_id"].isin(
                    set(pd.concat(picked)["contract_id"]) if picked else set()
                )]
                if len(remaining) > 0:
                    take = min(budget, len(remaining))
                    idx = rng.choice(len(remaining), size=take, replace=False)
                    picked.append(remaining.iloc[idx])
            result = pd.concat(picked) if picked else pd.DataFrame(columns=pool.columns)
            result = result.drop(columns=["_stratum"], errors="ignore")
        else:
            take = min(n, len(pool))
            idx = rng.choice(len(pool), size=take, replace=False)
            result = pool.iloc[idx]
        sampled_ids.update(result["contract_id"].tolist())
        return result

    # Stratum A: event=1, high_confidence_strict
    pool_a = df[(df["event"] == 1) & df["high_confidence_strict"]]
    frames.append(_pick(pool_a, 20))

    # Stratum B: event=1, ambiguous margin (low), multi-candidate
    pool_b = df[
        (df["event"] == 1)
        & (df["margin_bin"] == "low")
        & (df["n_candidates_for_source"] > 1)
    ]
    frames.append(_pick(pool_b, 20))

    # Stratum C: event=1, single candidate (not already in A/B)
    pool_c = df[(df["event"] == 1) & df["single_candidate_match"]]
    frames.append(_pick(pool_c, 15))

    # Stratum D: event=0, diverse categories + years
    pool_d = df[df["event"] == 0]
    frames.append(_pick(pool_d, 50, secondary_cols=["category_label", "start_year"]))

    # Stratum E: remaining event=1, medium/non-strict confidence
    pool_e = df[df["event"] == 1]
    remaining_target = TARGET_N - sum(len(f) for f in frames)
    frames.append(_pick(pool_e, remaining_target, secondary_cols=["category_label", "start_year"]))

    sample = pd.concat(frames, ignore_index=True)
    # Shuffle for unbiased audit ordering
    sample = sample.sample(frac=1, random_state=RANDOM_SEED).reset_index(drop=True)
    return sample


# ---------------------------------------------------------------------------
# 4. Enrich with candidate metadata
# ---------------------------------------------------------------------------
def enrich_candidates(sample: pd.DataFrame, candidates: pd.DataFrame) -> pd.DataFrame:
    sample = sample.copy()
    sample["_idweb"] = sample["contract_id"].str.replace("BOAMP:", "", regex=False)

    def _get(idweb: str, col: str):
        try:
            return candidates.at[idweb, col]
        except KeyError:
            return None

    sample["source_object_cand"] = sample["_idweb"].map(lambda x: _get(x, "src_objet"))
    sample["source_cpv_cand"] = sample["_idweb"].map(lambda x: _get(x, "src_cpv"))
    sample["cand_idweb"] = sample["_idweb"].map(lambda x: _get(x, "cand_idweb"))
    sample["candidate_date"] = sample["_idweb"].map(lambda x: _get(x, "cand_contract_start"))
    sample["candidate_cpv_cand"] = sample["_idweb"].map(lambda x: _get(x, "cand_cpv"))
    sample["candidate_object"] = sample["_idweb"].map(lambda x: _get(x, "cand_objet"))
    return sample


# ---------------------------------------------------------------------------
# 5. Enrich with boamp_full_clean (buyer names, URLs, objects)
# ---------------------------------------------------------------------------
def _idweb_from_contract_id(cid: str | float) -> str | None:
    if not isinstance(cid, str):
        return None
    return cid.replace("BOAMP:", "").strip()


def enrich_full_clean(
    sample: pd.DataFrame, full_clean: pd.DataFrame
) -> pd.DataFrame:
    sample = sample.copy()
    lookup = full_clean.set_index("idweb")

    def _buyer_name(idweb: str | None, buyer_key_fallback: str) -> str:
        if idweb and idweb in lookup.index:
            val = lookup.at[idweb, "nomacheteur"]
            if pd.notna(val) and str(val).strip():
                return str(val).strip()
        # Fallback: extract name from buyer_key like "NAME:some name"
        if isinstance(buyer_key_fallback, str) and buyer_key_fallback.startswith("NAME:"):
            return buyer_key_fallback[5:].strip().title()
        return str(buyer_key_fallback)

    def _url(idweb: str | None) -> str:
        if idweb and idweb in lookup.index:
            val = lookup.at[idweb, "url_avis"]
            if pd.notna(val) and str(val).strip():
                return str(val).strip()
        if idweb:
            return BOAMP_BASE_URL + idweb
        return ""

    def _obj(idweb: str | None, fallback: str | None) -> str:
        if pd.notna(fallback) and str(fallback).strip():
            return str(fallback).strip()
        if idweb and idweb in lookup.index:
            val = lookup.at[idweb, "objet"]
            if pd.notna(val):
                return str(val).strip()
        return ""

    src_idwebs = sample["contract_id"].map(_idweb_from_contract_id)
    cand_idwebs = sample["renewal_contract_id"].map(_idweb_from_contract_id)

    sample["source_buyer_name"] = [
        _buyer_name(iw, bk)
        for iw, bk in zip(src_idwebs, sample["buyer_key"])
    ]
    sample["candidate_buyer_name"] = [
        _buyer_name(iw, "")
        for iw in cand_idwebs
    ]
    sample["source_boamp_url"] = src_idwebs.map(_url)
    sample["candidate_boamp_url"] = cand_idwebs.map(_url)
    sample["source_object"] = [
        _obj(iw, so) for iw, so in zip(src_idwebs, sample["source_object_cand"])
    ]
    sample["source_cpv"] = sample["source_cpv_cand"].fillna(
        sample["cpv_div2"].astype(str)
    )
    sample["candidate_cpv"] = sample["candidate_cpv_cand"]
    return sample


# ---------------------------------------------------------------------------
# 6. Nearest later same-buyer notice for event=0
# ---------------------------------------------------------------------------
def add_nearest_later_notice(
    sample: pd.DataFrame, full_clean: pd.DataFrame
) -> pd.DataFrame:
    sample = sample.copy()
    sample["nearest_later_notice_id"] = None
    sample["nearest_later_notice_date"] = None

    event0_mask = sample["event"] == 0
    if not event0_mask.any():
        return sample

    # Build buyer→notices lookup
    fc = full_clean.dropna(subset=["buyer_key", "dateparution"]).copy()
    buyer_groups = {k: grp for k, grp in fc.groupby("buyer_key")}

    for idx, row in sample[event0_mask].iterrows():
        bk = row["buyer_key"]
        end = row["estimated_end_date"]
        if bk not in buyer_groups or pd.isna(end):
            continue
        grp = buyer_groups[bk]
        later = grp[grp["dateparution"] > end].sort_values("dateparution")
        if later.empty:
            continue
        best = later.iloc[0]
        sample.at[idx, "nearest_later_notice_id"] = str(best["idweb"])
        sample.at[idx, "nearest_later_notice_date"] = str(best["dateparution"].date())

    return sample


# ---------------------------------------------------------------------------
# 7. Build final audit table
# ---------------------------------------------------------------------------
def build_audit_table(sample: pd.DataFrame) -> pd.DataFrame:
    n = len(sample)
    sample = sample.reset_index(drop=True)
    audit_ids = [f"AUD-{i+1:03d}" for i in range(n)]

    audit = pd.DataFrame({
        "audit_id": audit_ids,
        "contract_id": sample["contract_id"],
        "event": sample["event"],
        "renewal_contract_id": sample["renewal_contract_id"],
        "source_buyer_name": sample.get("source_buyer_name", ""),
        "candidate_buyer_name": sample.get("candidate_buyer_name", ""),
        "source_object": sample.get("source_object", ""),
        "candidate_object": sample.get("candidate_object", ""),
        "source_cpv": sample.get("source_cpv", ""),
        "candidate_cpv": sample.get("candidate_cpv", ""),
        "category": sample["category_label"],
        "start_date": sample["start_date"].dt.strftime("%Y-%m-%d").where(
            sample["start_date"].notna(), other=""
        ),
        "estimated_end_date": sample["estimated_end_date"].dt.strftime("%Y-%m-%d").where(
            sample["estimated_end_date"].notna(), other=""
        ),
        "candidate_date": sample.get("candidate_date", ""),
        "observed_duration_months": sample["observed_duration_months"].round(2),
        "declared_duration_months": sample["declared_duration_months"].round(1),
        "composite_score": sample["composite_score"].round(4),
        "text_similarity": sample["text_similarity"].round(4),
        "cpv_score": sample["cpv_match_score"].round(4),
        "temporal_score": sample["temporal_score"].round(4),
        "score_margin": sample["score_margin"].round(4),
        "n_candidates": sample["n_candidates_for_source"],
        "high_confidence_strict": sample["high_confidence_strict"],
        "source_boamp_url": sample.get("source_boamp_url", ""),
        "candidate_boamp_url": sample.get("candidate_boamp_url", ""),
        "nearest_later_notice_id": sample.get("nearest_later_notice_id", ""),
        "nearest_later_notice_date": sample.get("nearest_later_notice_date", ""),
        # Manual judgment columns — intentionally left empty
        "manual_decision": "",
        "manual_error_type": "",
        "manual_notes": "",
        "boamp_source_record_checked": "",
        "boamp_candidate_record_checked": "",
    })
    return audit


# ---------------------------------------------------------------------------
# 8. Summary metrics
# ---------------------------------------------------------------------------
def build_validation_summary(df: pd.DataFrame) -> pd.DataFrame:
    ev1 = df[df["event"] == 1]
    rows = [
        ("total_contracts", len(df)),
        ("event_1_count", int((df["event"] == 1).sum())),
        ("event_0_count", int((df["event"] == 0).sum())),
        ("event_rate", round((df["event"] == 1).mean(), 4)),
        ("high_confidence_strict_count", int(df["high_confidence_strict"].sum())),
        ("high_confidence_strict_pct", round(df["high_confidence_strict"].mean(), 4)),
        ("median_composite_score_event1",
         round(ev1["composite_score"].median(), 4) if len(ev1) > 0 else None),
        ("median_score_margin_event1",
         round(ev1["score_margin"].median(), 4) if len(ev1) > 0 else None),
        ("single_candidate_count", int(df["single_candidate_match"].sum())),
        ("single_candidate_pct", round(df["single_candidate_match"].mean(), 4)),
        ("composite_score_p25_event1",
         round(ev1["composite_score"].quantile(0.25), 4) if len(ev1) > 0 else None),
        ("composite_score_p75_event1",
         round(ev1["composite_score"].quantile(0.75), 4) if len(ev1) > 0 else None),
    ]
    return pd.DataFrame(rows, columns=["metric", "value"])


def build_threshold_sensitivity(df: pd.DataFrame) -> pd.DataFrame:
    # composite_score is only populated for event=1 rows (no match → no score).
    # The threshold asks: "how many of the flagged event=1 contracts meet this quality bar?"
    # effective_event_rate = retained matches / total contracts (useful for model calibration).
    rows = []
    total = len(df)
    n_event1 = int((df["event"] == 1).sum())
    for thr in SCORE_THRESHOLDS:
        above = df[(df["event"] == 1) & (df["composite_score"] >= thr)]
        n_above = len(above)
        pct_of_event1 = round(n_above / max(1, n_event1), 4)
        effective_event_rate = round(n_above / total, 4)
        rows.append({
            "threshold": thr,
            "event1_retained": n_above,
            "pct_of_all_event1_retained": pct_of_event1,
            "effective_event_rate_full_dataset": effective_event_rate,
        })
    return pd.DataFrame(rows)


def build_event_bias_summary(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["start_year"] = df["start_date"].dt.year
    grp = df.groupby("start_year", dropna=False).agg(
        total=("event", "count"),
        event1_count=("event", "sum"),
        mean_composite_score=("composite_score", "mean"),
    ).reset_index()
    grp["event_rate"] = (grp["event1_count"] / grp["total"]).round(4)
    grp["mean_composite_score"] = grp["mean_composite_score"].round(4)
    return grp[["start_year", "total", "event1_count", "event_rate", "mean_composite_score"]]


def build_event_rate_by_category(df: pd.DataFrame) -> pd.DataFrame:
    grp = df.groupby("category_label", dropna=False).agg(
        total=("event", "count"),
        event1_count=("event", "sum"),
        mean_composite_score=("composite_score", "mean"),
    ).reset_index()
    grp["event_rate"] = (grp["event1_count"] / grp["total"]).round(4)
    grp["mean_composite_score"] = grp["mean_composite_score"].round(4)
    return grp.sort_values("total", ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# 9. Write Excel workbook
# ---------------------------------------------------------------------------
def write_excel(
    audit: pd.DataFrame,
    summary: pd.DataFrame,
    threshold: pd.DataFrame,
    bias: pd.DataFrame,
    category: pd.DataFrame,
    out_path: Path,
) -> None:
    readme_rows = [
        ("Workbook", "BOAMP Proxy Event Manual Validation"),
        ("Purpose",
         "150-row stratified audit sample for manual review of the renewal proxy event."),
        ("event=1",
         "Algorithm found an identifiable BOAMP successor notice (proxy renewal event)."),
        ("event=0",
         "No identifiable BOAMP successor was found (censored or true non-renewal)."),
        ("Manual_Audit_Sample",
         "150-row audit table. Fill manual_decision and manual_error_type columns only."),
        ("Validation_Metrics",
         "Population-level statistics for the full survival dataset (all 1100 rows)."),
        ("Threshold_Sensitivity",
         "Event=1 retention and implied event rate at different composite_score thresholds."),
        ("Event_Bias_Check",
         "Event rate and composite score by calendar year — checks for temporal bias."),
        ("Category_Event_Rates",
         "Event rate by technology category — checks for category-level bias."),
        ("manual_decision allowed values",
         " | ".join(MANUAL_DECISION_VALUES)),
        ("manual_error_type allowed values",
         " | ".join(MANUAL_ERROR_TYPE_VALUES)),
        ("Random seed", "42"),
        ("Generated by", "event_validation/scripts/build_validation_sample.py"),
    ]
    readme_df = pd.DataFrame(readme_rows, columns=["Key", "Description"])

    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            sheets = {
                "README": readme_df,
                "Manual_Audit_Sample": audit,
                "Validation_Metrics": summary,
                "Threshold_Sensitivity": threshold,
                "Event_Bias_Check": bias,
                "Category_Event_Rates": category,
            }
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply formatting
            wb = writer.book
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="D9E1F2")

            for sheet_name in sheets:
                ws = wb[sheet_name]
                # Bold + color header row
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                # Freeze top row
                ws.freeze_panes = "A2"
                # Auto-size columns (cap at 50)
                for col_idx, col_cells in enumerate(ws.columns, 1):
                    max_len = max(
                        (len(str(c.value)) for c in col_cells if c.value is not None),
                        default=8,
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(
                        max_len + 2, 50
                    )
    except ImportError:
        print("openpyxl not found — writing Excel without formatting.")
        with pd.ExcelWriter(out_path) as writer:
            readme_df.to_excel(writer, sheet_name="README", index=False)
            audit.to_excel(writer, sheet_name="Manual_Audit_Sample", index=False)
            summary.to_excel(writer, sheet_name="Validation_Metrics", index=False)
            threshold.to_excel(writer, sheet_name="Threshold_Sensitivity", index=False)
            bias.to_excel(writer, sheet_name="Event_Bias_Check", index=False)
            category.to_excel(writer, sheet_name="Category_Event_Rates", index=False)

    print(f"  Excel written: {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    rng = np.random.default_rng(RANDOM_SEED)

    print("Loading survival dataset …")
    survival = load_survival()
    print(f"  {len(survival)} rows | event=1: {(survival['event']==1).sum()} | event=0: {(survival['event']==0).sum()}")

    print("Loading candidates dataset …")
    candidates = load_candidates()
    print(f"  {len(candidates)} unique source contracts in candidates")

    print("Loading boamp_full_clean …")
    full_clean = load_full_clean()
    print(f"  {len(full_clean)} rows in full_clean")

    print("Engineering strata features …")
    survival = add_strata_features(survival)

    print("Drawing stratified 150-row sample …")
    sample = stratified_sample(survival, rng)
    print(f"  Sampled {len(sample)} rows | event=1: {(sample['event']==1).sum()} | event=0: {(sample['event']==0).sum()}")

    print("Enriching with candidate metadata …")
    sample = enrich_candidates(sample, candidates)

    print("Enriching with boamp_full_clean …")
    sample = enrich_full_clean(sample, full_clean)

    print("Finding nearest later same-buyer notices for event=0 …")
    sample = add_nearest_later_notice(sample, full_clean)

    print("Building audit table …")
    audit = build_audit_table(sample)

    print("Building summary tables …")
    summary_metrics = build_validation_summary(survival)
    threshold_sens = build_threshold_sensitivity(survival)
    bias_summary = build_event_bias_summary(survival)
    cat_rates = build_event_rate_by_category(survival)

    print("Writing CSV outputs …")
    audit.to_csv(OUT_DIR / "manual_validation_sample.csv", index=False)
    print(f"  manual_validation_sample.csv ({len(audit)} rows)")
    summary_metrics.to_csv(OUT_DIR / "validation_summary_metrics.csv", index=False)
    print(f"  validation_summary_metrics.csv")
    threshold_sens.to_csv(OUT_DIR / "threshold_sensitivity.csv", index=False)
    print(f"  threshold_sensitivity.csv")
    bias_summary.to_csv(OUT_DIR / "event_bias_summary.csv", index=False)
    print(f"  event_bias_summary.csv")
    cat_rates.to_csv(OUT_DIR / "event_rate_by_category.csv", index=False)
    print(f"  event_rate_by_category.csv")

    print("Writing Excel workbook …")
    write_excel(
        audit, summary_metrics, threshold_sens, bias_summary, cat_rates,
        OUT_DIR / "boamp_event_validation_audit.xlsx",
    )

    # Verification checks
    print("\n--- Verification ---")
    print(f"  Audit rows: {len(audit)}")
    print(f"  Event distribution: {audit['event'].value_counts().to_dict()}")
    manual_cols = ["manual_decision", "manual_error_type", "manual_notes",
                   "boamp_source_record_checked", "boamp_candidate_record_checked"]
    all_empty = all(audit[c].fillna("").eq("").all() for c in manual_cols)
    print(f"  Manual columns all empty: {all_empty}")
    print(f"  Category coverage: {audit['category'].nunique()} categories")
    print(f"  Year coverage: {sorted(pd.to_datetime(audit['start_date'], errors='coerce').dt.year.dropna().unique().astype(int).tolist())}")
    print(f"\nAll outputs written to: {OUT_DIR}")


if __name__ == "__main__":
    main()
