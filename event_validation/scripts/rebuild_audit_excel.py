"""
Rebuild boamp_event_validation_audit.xlsx with:
  - Dropdowns for manual_decision
  - Auto-formula for manual_error_type (fills itself from manual_decision)
  - Dropdown (yes/no) for boamp_*_checked
  - Clickable hyperlinks for BOAMP URLs
  - Color-coded rows (event=1 blue, event=0 yellow)
  - Highlighted manual columns
  - Frozen header + first 3 columns
  - A focused "Working" sheet with only judgment-relevant columns
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from pathlib import Path

ROOT  = Path(__file__).resolve().parents[2]
OUT   = ROOT / "event_validation" / "outputs" / "boamp_event_validation_audit.xlsx"
CSV   = ROOT / "event_validation" / "outputs" / "manual_validation_sample.csv"

# ── Load existing workbook (keep non-audit sheets) ─────────────────────────
wb_old = load_workbook(OUT)
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── colours ────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy header
EV1_FILL   = PatternFill("solid", fgColor="DDEEFF")   # light blue  event=1
EV0_FILL   = PatternFill("solid", fgColor="FFFDE7")   # light yellow event=0
MAN_FILL   = PatternFill("solid", fgColor="E8F5E9")   # light green manual cols
DONE_FILL  = PatternFill("solid", fgColor="C8E6C9")   # green  credible_renewal
FP_FILL    = PatternFill("solid", fgColor="FFCDD2")   # red    false positive
DOUBT_FILL = PatternFill("solid", fgColor="FFF9C4")   # yellow doubtful

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(size=9)
LINK_FONT  = Font(size=9, color="1155CC", underline="single")
WRAP       = Alignment(wrap_text=True, vertical="top")
TOP        = Alignment(vertical="top")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── dropdown values ─────────────────────────────────────────────────────────
DECISION_OPTIONS = (
    '"credible_renewal,doubtful_but_possible,'
    'not_credible_false_positive,'
    'plausible_censored,'
    'missed_renewal_false_negative,'
    'impossible_to_judge"'
)
CHECKED_OPTIONS = '"yes,no"'

# ── manual_error_type formula (col letter filled in later) ──────────────────
def error_type_formula(dec_col: str, row: int) -> str:
    c = f"{dec_col}{row}"
    return (
        f'=IF({c}="credible_renewal","true_positive",'
        f'IF({c}="not_credible_false_positive","false_positive",'
        f'IF({c}="plausible_censored","true_negative_or_plausible_censored",'
        f'IF({c}="missed_renewal_false_negative","false_negative",'
        f'IF(OR({c}="doubtful_but_possible",{c}="impossible_to_judge"),'
        f'"not_applicable","")))))'
    )

# ── helper: set header row ──────────────────────────────────────────────────
def write_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 30

# ── helper: add decision dropdown + formula + checked dropdowns ─────────────
def add_validation(ws, n_rows, dec_col_idx, err_col_idx,
                   src_chk_col_idx, cnd_chk_col_idx):
    dec_col = get_column_letter(dec_col_idx)
    err_col = get_column_letter(err_col_idx)

    # Dropdown: manual_decision
    dv_dec = DataValidation(
        type="list", formula1=DECISION_OPTIONS,
        allow_blank=True, showErrorMessage=False,
        showInputMessage=True,
        promptTitle="Choose decision",
        prompt="Pick one value from the list"
    )
    ws.add_data_validation(dv_dec)
    dv_dec.sqref = f"{dec_col}2:{dec_col}{n_rows + 1}"

    # Dropdown: boamp_source_record_checked
    dv_src = DataValidation(type="list", formula1=CHECKED_OPTIONS,
                            allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_src)
    dv_src.sqref = f"{get_column_letter(src_chk_col_idx)}2:{get_column_letter(src_chk_col_idx)}{n_rows+1}"

    # Dropdown: boamp_candidate_record_checked
    dv_cnd = DataValidation(type="list", formula1=CHECKED_OPTIONS,
                            allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cnd)
    dv_cnd.sqref = f"{get_column_letter(cnd_chk_col_idx)}2:{get_column_letter(cnd_chk_col_idx)}{n_rows+1}"

    # Formula: manual_error_type (auto-fills)
    for r in range(2, n_rows + 2):
        cell = ws.cell(r, err_col_idx)
        cell.value = error_type_formula(dec_col, r)
        cell.font  = Font(size=9, italic=True, color="555555")

# ── helper: row fill by event value ─────────────────────────────────────────
def row_fill(event_val):
    return EV1_FILL if event_val == 1 else EV0_FILL

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Working (focused view for annotation)
# ════════════════════════════════════════════════════════════════════════════
df = pd.read_csv(CSV)

WORK_COLS = [
    "audit_id", "event",
    "source_buyer_name", "candidate_buyer_name",
    "source_object", "candidate_object",
    "source_cpv", "candidate_cpv",
    "start_date", "estimated_end_date", "candidate_date",
    "composite_score", "text_similarity", "cpv_score",
    "temporal_score", "score_margin",
    "source_boamp_url", "candidate_boamp_url",
    "nearest_later_notice_id", "nearest_later_notice_date",
    "manual_decision", "manual_error_type", "manual_notes",
    "boamp_source_record_checked", "boamp_candidate_record_checked",
]

ws_work = wb.create_sheet("Working")
write_header(ws_work, WORK_COLS)

# column index map for Working sheet
wi = {col: i + 1 for i, col in enumerate(WORK_COLS)}

URL_COLS_WORK = {"source_boamp_url", "candidate_boamp_url"}
MAN_COLS_WORK = {"manual_decision", "manual_error_type", "manual_notes",
                 "boamp_source_record_checked", "boamp_candidate_record_checked"}

for r, row in enumerate(df.itertuples(index=False), 2):
    ev = getattr(row, "event", 0)
    fill = row_fill(ev)

    for col in WORK_COLS:
        if col in ("manual_decision", "manual_error_type",
                   "boamp_source_record_checked", "boamp_candidate_record_checked"):
            continue  # handled separately

        val = getattr(row, col, "")
        val = "" if str(val) == "nan" else val
        c_idx = wi[col]
        cell = ws_work.cell(r, c_idx, val)
        cell.border = BORDER
        cell.fill = MAN_FILL if col in MAN_COLS_WORK else fill

        if col in URL_COLS_WORK and val:
            cell.hyperlink = str(val)
            cell.value = str(val).replace("https://www.boamp.fr/pages/avis/?q=idweb:", "")
            cell.font = LINK_FONT
        elif col in ("source_object", "candidate_object"):
            cell.alignment = WRAP
            cell.font = BODY_FONT
        elif col == "manual_notes":
            cell.alignment = WRAP
            cell.font = BODY_FONT
        else:
            cell.alignment = TOP
            cell.font = BODY_FONT

add_validation(
    ws_work, len(df),
    dec_col_idx=wi["manual_decision"],
    err_col_idx=wi["manual_error_type"],
    src_chk_col_idx=wi["boamp_source_record_checked"],
    cnd_chk_col_idx=wi["boamp_candidate_record_checked"],
)

# column widths — Working
col_widths_work = {
    "audit_id": 9, "event": 6,
    "source_buyer_name": 28, "candidate_buyer_name": 28,
    "source_object": 42, "candidate_object": 42,
    "source_cpv": 12, "candidate_cpv": 12,
    "start_date": 12, "estimated_end_date": 14, "candidate_date": 14,
    "composite_score": 12, "text_similarity": 12,
    "cpv_score": 10, "temporal_score": 12, "score_margin": 11,
    "source_boamp_url": 16, "candidate_boamp_url": 16,
    "nearest_later_notice_id": 22, "nearest_later_notice_date": 18,
    "manual_decision": 26, "manual_error_type": 36,
    "manual_notes": 36,
    "boamp_source_record_checked": 14, "boamp_candidate_record_checked": 16,
}
for col, width in col_widths_work.items():
    ws_work.column_dimensions[get_column_letter(wi[col])].width = width

ws_work.freeze_panes = "C2"  # freeze audit_id + event columns + header row
ws_work.row_dimensions[1].height = 38

# conditional formatting on manual_decision column (Working)
dec_col_w = get_column_letter(wi["manual_decision"])
rng = f"{dec_col_w}2:{dec_col_w}{len(df)+1}"
ws_work.conditional_formatting.add(rng, FormulaRule(
    formula=[f'{dec_col_w}2="credible_renewal"'],
    fill=DONE_FILL, font=Font(size=9, bold=True, color="1B5E20")
))
ws_work.conditional_formatting.add(rng, FormulaRule(
    formula=[f'{dec_col_w}2="not_credible_false_positive"'],
    fill=FP_FILL, font=Font(size=9, bold=True, color="B71C1C")
))
ws_work.conditional_formatting.add(rng, FormulaRule(
    formula=[f'{dec_col_w}2="doubtful_but_possible"'],
    fill=DOUBT_FILL, font=Font(size=9, color="5D4037")
))

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Manual_Audit_Sample (full 32-col reference sheet)
# ════════════════════════════════════════════════════════════════════════════
ALL_COLS = df.columns.tolist()
ws_full = wb.create_sheet("Manual_Audit_Sample")
write_header(ws_full, ALL_COLS)

ai = {col: i + 1 for i, col in enumerate(ALL_COLS)}
URL_COLS_FULL = {"source_boamp_url", "candidate_boamp_url"}
MAN_COLS_FULL = {"manual_decision", "manual_error_type", "manual_notes",
                 "boamp_source_record_checked", "boamp_candidate_record_checked"}

for r, row in enumerate(df.itertuples(index=False), 2):
    ev = getattr(row, "event", 0)
    fill = row_fill(ev)
    for col in ALL_COLS:
        if col in ("manual_decision", "manual_error_type",
                   "boamp_source_record_checked", "boamp_candidate_record_checked"):
            continue
        val = getattr(row, col, "")
        val = "" if str(val) == "nan" else val
        c_idx = ai[col]
        cell = ws_full.cell(r, c_idx, val)
        cell.border = BORDER
        cell.fill = MAN_FILL if col in MAN_COLS_FULL else fill
        if col in URL_COLS_FULL and val:
            cell.hyperlink = str(val)
            cell.value = str(val).replace("https://www.boamp.fr/pages/avis/?q=idweb:", "")
            cell.font = LINK_FONT
        else:
            cell.alignment = TOP
            cell.font = BODY_FONT

add_validation(
    ws_full, len(df),
    dec_col_idx=ai["manual_decision"],
    err_col_idx=ai["manual_error_type"],
    src_chk_col_idx=ai["boamp_source_record_checked"],
    cnd_chk_col_idx=ai["boamp_candidate_record_checked"],
)

# column widths — Full sheet
col_widths_full = {c: 14 for c in ALL_COLS}
col_widths_full.update({
    "audit_id": 9, "contract_id": 18, "event": 6,
    "source_buyer_name": 26, "candidate_buyer_name": 26,
    "source_object": 36, "candidate_object": 36,
    "source_boamp_url": 16, "candidate_boamp_url": 16,
    "manual_decision": 26, "manual_error_type": 36,
    "manual_notes": 36,
})
for col, width in col_widths_full.items():
    ws_full.column_dimensions[get_column_letter(ai[col])].width = width

ws_full.freeze_panes = "A2"

# ════════════════════════════════════════════════════════════════════════════
# COPY remaining sheets from original workbook
# ════════════════════════════════════════════════════════════════════════════
KEEP_SHEETS = ["README", "Validation_Metrics", "Threshold_Sensitivity",
               "Event_Bias_Check", "Category_Event_Rates", "BOAMP_Raw_Records"]

for name in KEEP_SHEETS:
    if name not in wb_old.sheetnames:
        continue
    ws_src = wb_old[name]
    ws_dst = wb.create_sheet(name)
    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font      = cell.font.copy()
                new_cell.fill      = cell.fill.copy()
                new_cell.border    = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()

# ── Save ───────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Saved → {OUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"\nWorking sheet: {len(df)} data rows, {len(WORK_COLS)} columns")
print("  → manual_decision  : dropdown (6 options)")
print("  → manual_error_type: auto-formula (fills from manual_decision)")
print("  → boamp_*_checked  : dropdown (yes / no)")
print("  → BOAMP URLs       : clickable hyperlinks")
